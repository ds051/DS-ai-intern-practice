import os
import time
import json
import requests
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill

# ================= 1. 安全加载配置 =================
load_dotenv()  # 自动加载 .env 文件中的变量
API_KEY = os.getenv("DEEPSEEK_API_KEY")

if not API_KEY:
    raise ValueError("❌ 未找到 DEEPSEEK_API_KEY！请在当前目录下创建 .env 文件并填入密钥。")

# 配置区
INPUT_FOLDER = "./"      # 待处理的txt文件夹
OUTPUT_EXCEL = "处理结果汇总.xlsx"
MODEL_NAME = "deepseek-chat"      # 使用 deepseek-chat (V3)
MAX_RETRIES = 3                   # 失败重试次数

# ================= 2. 核心 AI 处理函数 =================
def call_deepseek_api(content):
    """
    调用 DeepSeek API 生成摘要和关键词
    返回: (summary_str, keywords_str)
    """
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {API_KEY}"
    }
    
    # 构造 Prompt，强制要求 JSON 格式返回以便程序解析
    prompt = f"""请阅读以下文本，完成两个任务：
1. 生成一段100字以内的摘要。
2. 提取3个最核心的关键词。

请严格按照以下 JSON 格式返回，不要包含任何多余的解释文字：
{{
    "summary": "这里是摘要内容...",
    "keywords": ["关键词1", "关键词2", "关键词3"]
}}

文本内容如下：
{content[:3000]} 
""" 
    # 注：这里限制了3000字以防超出上下文窗口，可根据需要调整

    payload = {
        "model": MODEL_NAME,
        "messages": [{"role": "user", "content": prompt}],
        "temperature": 0.3  # 较低的温度让输出更稳定
    }

    response = requests.post(
        "https://api.deepseek.com/v1/chat/completions",
        headers=headers,
        json=payload,
        timeout=60
    )
    
    if response.status_code != 200:
        raise Exception(f"API 请求失败: {response.status_code} - {response.text}")

    result_json = response.json()
    content_str = result_json['choices'][0]['message']['content']
    
    # 尝试解析 JSON
    try:
        # 清理可能存在的 markdown 代码块标记 ```json ... ```
        clean_str = content_str.replace("```json", "").replace("```", "").strip()
        data = json.loads(clean_str)
        return data.get("summary", "无摘要"), ", ".join(data.get("keywords", []))
    except json.JSONDecodeError:
        # 如果模型偶尔没按 JSON 返回，做简单的兜底处理
        return content_str.strip(), "解析失败"

# ================= 3. 批量处理主逻辑 =================
def process_folder():
    results = []
    failed_files = []
    
    # 获取所有 txt 文件
    files = list(Path(INPUT_FOLDER).glob("*.txt"))
    print(f"🔍 发现 {len(files)} 个文件，开始处理...")

    for file_path in files:
        file_name = file_path.name
        summary, keywords = "", ""
        success = False
        
        # --- 重试机制 ---
        for attempt in range(MAX_RETRIES):
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read().strip()
                
                if not content:
                    raise ValueError("文件为空")
                    
                summary, keywords = call_deepseek_api(content)
                success = True
                print(f"✅ 成功: {file_name}")
                break  # 成功则跳出重试循环
                
            except Exception as e:
                error_msg = str(e)
                print(f"⚠️ 失败 ({attempt+1}/{MAX_RETRIES}): {file_name} -> {error_msg}")
                if attempt < MAX_RETRIES - 1:
                    time.sleep(2)  # 等待几秒再重试
        
        if success:
            results.append({
                "文件名": file_name,
                "摘要": summary,
                "关键词": keywords,
                "状态": "成功"
            })
        else:
            failed_files.append({"文件名": file_name, "错误原因": error_msg})
            results.append({
                "文件名": file_name,
                "摘要": "处理失败",
                "关键词": "-",
                "状态": f"失败: {error_msg[:50]}"
            })

    return results, failed_files

# ================= 4. 导出美化 Excel =================
def save_to_excel(data_list):
    df = pd.DataFrame(data_list)
    
    # 使用 Pandas 写入 Excel
    with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='处理结果')
        
        # 获取 workbook 和 worksheet 对象进行格式化
        wb = writer.book
        ws = writer.sheets['处理结果']
        
        # 设置表头样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            
            for cell in col_cells:
                # 设置自动换行
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # 计算最大宽度
                if cell.value:
                    length = len(str(cell.value))
                    # 中文字符算2个宽度
                    length += sum(1 for c in str(cell.value) if '\u4e00' <= c <= '\u9fff')
                    max_length = max(max_length, length)
                
                # 给第一行（表头）加样式
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')

            # 调整列宽 (限制最大宽度防止太宽，最小宽度防止太窄)
            adjusted_width = min(max(max_length + 2, 10), 80) 
            ws.column_dimensions[col_letter].width = adjusted_width
            
        # 特别加宽“摘要”列
        ws.column_dimensions['B'].width = 60 

    print(f"\n🎉 处理完成！结果已保存至: {os.path.abspath(OUTPUT_EXCEL)}")

# ================= 5. 运行入口 =================
if __name__ == "__main__":
    if not os.path.exists(INPUT_FOLDER):
        print(f"❌ 找不到文件夹: {INPUT_FOLDER}，请检查路径。")
    else:
        results, failed = process_folder()
        save_to_excel(results)
        
        if failed:
            print("\n⚠️ 以下文件最终处理失败:")
            for f in failed:
                print(f"   - {f['文件名']}: {f['错误原因']}")